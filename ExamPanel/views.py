from django.shortcuts import render, redirect
from .forms import *
import itertools
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import *
from datetime import date,datetime,timedelta
from Frontdesk.models import *
from Institute.views import *
from django.core.urlresolvers import resolve
from django.shortcuts import HttpResponseRedirect
from django.core.urlresolvers import reverse
from Frontdesk.views import *
import csv
import openpyxl

def Institute_data(u = None):
    if u:
        ins_data = Institite_profile.objects.filter(id=int(u.first_name)).first()
        return ins_data
    ins_data = Institite_profile.objects.filter().last()
    return ins_data

def is_admin(user):
    if user.last_name == "Admin":
        return True
    return False


def is_Examinar(user):
    if user.last_name == "Exam":
        return True
    return False



def Send_msg_for_new_test_series_asign(sc):
    setting_data = Auto_SMS_Settings_data.objects.filter().first()
    if setting_data:
        if setting_data.new_ts_asign:
            sms = setting_data.sms_new_ts_asign.sms
            SplitSMS = sms.split("[[")
            final = SplitSMS[0]

            ins_data = Institite_profile.objects.filter().first()

            for i in range(1, len(SplitSMS)):

                msg = SplitSMS[i]
                z = msg.split("]]")

                w = z.pop(0)
                add = ''
                if w == "StudentName":
                    add = sc.student.name
                if w == "Course":
                    add = sc.course.name
                if w == "FeesPackageAmount":
                    add = sc.total_fee
                if w == "TotalReceivedAmount":
                    add = sc.paid_fee
                if w == "TotalReceivableAmount":
                    add = sc.remain_fee
                if w == "InvoiceNo":
                    add = sc.examinvoicegenerate.last.invoice
                if w == "PaymentMode":
                    add = sc.examinvoicegenerate.last.mode
                if w == "BranchName":
                    add = ins_data.name
                if w == "BranchEmail":
                    add = ins_data.email

                if w == "BranchMobileNo":
                    add = ins_data.office_mob

                final += " " + add + " " + z[0]

            smsnum = ceil(len(final)/160)
            want_sms = smsnum * 1
            Now = NowSMS.objects.filter().first()
            ava = int(Now.num)
            if ava < want_sms:
                return redirect("Master:sms_fail")
            SendSMS(sc.student.mobile, final, setting_data.sms_new_ts_asign.sender_id)


def CreateNewQuestion(que):
    sque = StudentQuestions.objects.create(section=que.section, Type1=que.Type1, Questions=que.Questions,
                                           Option_1=que.Option_1, Option_2=que.Option_2, Option_3=que.Option_3,
                                           Option_4=que.Option_4, TrueAns=que.TrueAns, marks1=que.marks1,
                                           minus1=que.minus1,
                                           explain1=que.explain1, Type2=que.Type2, Question=que.Question,
                                           Option1=que.Option1,
                                           Option2=que.Option2, Option3=que.Option3,
                                           Option4=que.Option4, Answer1=que.Answer1, Answer2=que.Answer2,
                                           Answer3=que.Answer3, Answer4=que.Answer3,
                                           marks2=que.marks2, minus2=que.minus2, explain2=que.explain2, Type3=que.Type3,
                                           Que=que.Que, Answer31=que.Answer31,
                                           Answer32=que.Answer32, marks3=que.marks3, minus3=que.minus3,
                                           explain3=que.explain3, Type4=que.Type4, Ques_tion=que.Ques_tion,
                                           TrueAns1=que.TrueAns1, marks4=que.marks4, minus4=que.minus4,
                                           explain4=que.explain4)
    return sque




'''Quiz Package View Funcations'''
def CreateQuizPackage(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = QuizPackage.objects.all().order_by('-id')
    form = CreateQuizPackageForm()
    if request.method == "POST":
        form = CreateQuizPackageForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:create_quiz_package")
    return render(request, "exam/create_quiz_package.html", {"form":form, 'packages':data
                                                             ,"admin":admin_show()
                                                         ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditQuizPackage(request, eid, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if what == "Edit":
        data = QuizPackage.objects.filter(id = eid).first()
        form = CreateQuizPackageForm(request.POST or None, request.FILES or None, instance=data)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:create_quiz_package")

        return render(request, "exam/edit_quiz_package.html", {"form":form,"admin":admin_show()
                                                         ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})
    if what == "Delete":
        data = QuizPackage.objects.filter(id=eid).first()
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))



def CreateQuiz(request, cour, pack, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = Quiz.objects.all().order_by('-id')
    course = Sub_course.objects.all()
    pck = []
    cou = 0
    if cour != "0":
        cou = Sub_course.objects.filter(id=cour).first()
        pck = QuizPackage.objects.filter(course=cou)
    if request.method == "POST":
        if pack == '0':
            return redirect("ExamPanel:create_quiz", cour, '0')
        pck = QuizPackage.objects.filter(id = pack).first()
        name = request.POST["name"]
        duration = request.POST["duration"]
        quiz = Quiz.objects.create(course = cou, name = name, duration = duration)
        Quiz_QuizPackage.objects.create(package = pck, quiz = quiz)
        if what == "New":
            return redirect("ExamPanel:create_quiz", '0', '0', 'New')
        return redirect(request.META.get('HTTP_REFERER'))

    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'package': pck, 'ipack': int(pack), 'pack': pack,
         'subpackages':data, "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
        }
    return render(request, "exam/create_quiz.html", context)



def EditQuiz(request, eid, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if what == "Edit":
        data = Quiz.objects.filter(id = eid).first()
        if request.method == "POST":
            name = request.POST["name"]
            duration = request.POST["duration"]
            data.name = name
            data.duration = duration
            data.save()
            return redirect("ExamPanel:create_quiz", '0', '0', 'New')

        return render(request, "exam/edit_quiz.html", {"data":data,"admin":admin_show()
                                                         ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})
    if what == "Delete":
        data = Quiz.objects.filter(id=eid).first()
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))



def EditStudentQuiz(request, eid, what, sid, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if what == "Edit":
        data = StudentQuiz.objects.filter(id = eid).first()
        if request.method == "POST":
            name = request.POST["name"]
            duration = request.POST["duration"]
            data.name = name
            data.duration = duration
            data.save()
            if option == "Institute":
                return redirect("ExamPanel:institute_students_view", 'View', sid)
            if option == "External":
                return redirect("ExamPanel:other_students_view", 'View', sid)

        return render(request, "exam/edit_quiz.html", {"data":data,"admin":admin_show()
                                                         ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})
    if what == "Delete":
        data = StudentQuiz.objects.filter(id=eid).first()
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))






def QuizList(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    pack = QuizPackage.objects.filter(id = eid).first()
    sub = Quiz_QuizPackage.objects.filter(package = pack)
    courses = Sub_course.objects.all()
    if request.method == "POST":
        d = request.POST.getlist('checks')
        for i in d:
            n = 0
            for j in sub:
                if i == str(j.quiz.id):
                    n = n+1
            if n == 0:
                s = Quiz.objects.filter(id = i).first()
                Quiz_QuizPackage.objects.create(package=pack, quiz = s)
        return redirect("ExamPanel:add_quiz", eid)

    context = {
        'subpackages':sub,
        'courses':courses,"ins_data":Institute_data(request.user), "adminprofile":Admin_data(), "pack":pack
    }
    return render(request, "exam/quiz_list.html", context)


def AddQuestion_In_Quiz(request, q_id, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(q_id) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    sections = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        sections = Question_Section.objects.filter(course = cour)
        if sid != "0":
            section = Question_Section.objects.filter(id=sid).first()
            all_ques_tion = MultipleChoiceQuestions.objects.filter(section = section)

            quiz = Quiz.objects.filter(id=q_id).first()
            pre_que = Quiz_Questions.objects.filter(quiz=quiz)


            for i in all_ques_tion:
                n = 0
                for j in pre_que:
                    if i.id == j.Question.id:
                        n += 1
                if n == 0:
                    ques_tion.append(i)

            page = request.GET.get('page', 1)
            num = int(pno) * (int(page)-1)
            paginator = Paginator(ques_tion, int(pno))
            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)

            if request.method == "POST":
                qu = request.POST.getlist('que')
                sec = Quiz.objects.filter(id=q_id).first()
                print(sec)
                for i in qu:
                    que = MultipleChoiceQuestions.objects.filter(id = i).first()
                    print(que)
                    if que:
                        Quiz_Questions.objects.create(quiz = sec, Question = que)
                return redirect("ExamPanel:add_questions_in_quiz", q_id, cid, sid, pno)

    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'sections':sections, 'sec':sid, "isec":int(sid),
        'questions': data, 'users': data, "complete_url":complete_url,
        'pno':pno, 'cid':cid, 'num':num, 's_id':q_id,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/add_question_in_quiz.html", context)


def AddQuestion_In_Student_Quiz(request, q_id, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(q_id) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    sections = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        sections = Question_Section.objects.filter(course = cour)
        if sid != "0":
            section = Question_Section.objects.filter(id=sid).first()
            all_ques_tion = MultipleChoiceQuestions.objects.filter(section = section)

            quiz = StudentQuiz.objects.filter(id=q_id).first()
            pre_que = StudentQuiz_Questions.objects.filter(quiz=quiz)

            for i in all_ques_tion:
                n = 0
                for j in pre_que:
                    if i.id == j.Question.id:
                        n += 1
                if n == 0:
                    ques_tion.append(i)

            page = request.GET.get('page', 1)
            num = int(pno) * (int(page)-1)
            paginator = Paginator(ques_tion, int(pno))
            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)

            if request.method == "POST":
                qu = request.POST.getlist('que')
                sec = StudentQuiz.objects.filter(id=q_id).first()
                for i in qu:
                    que = MultipleChoiceQuestions.objects.filter(id = i).first()
                    if que:
                        StudentQuiz_Questions.objects.create(quiz = sec, Question = que)
                return redirect("ExamPanel:add_questions_in_student_quiz", q_id, cid, sid, pno)

    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'sections':sections, 'sec':sid, "isec":int(sid),
        'questions': data, 'users': data, "complete_url":complete_url,
        'pno':pno, 'cid':cid, 'num':num, 's_id':q_id,"admin":admin_show(),
        "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/add_question_in_student_quiz.html", context)






def QuizQuestionList(request, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if sid != "0":
        sec = Quiz.objects.filter(id = sid).first()
        all_eq = Quiz_Questions.objects.filter(quiz = sec)
        for i in all_eq:
            ques_tion.append(i)

        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'questions': data, 'users': data,
        'pno':pno, 'num':num, 'cid':sid,
        'complete_url':complete_url,"admin":admin_show(),
        "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/QuizQuestionList.html", context)

def StudentQuizQuestionList(request, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if sid != "0":
        sec = StudentQuiz.objects.filter(id = sid).first()
        all_eq = StudentQuiz_Questions.objects.filter(quiz = sec)
        for i in all_eq:
            ques_tion.append(i)

        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'questions': data, 'users': data,
        'pno':pno, 'num':num, 'cid':sid,
        'complete_url':complete_url,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/StudentQuizQuestionList.html", context)



######## UpLoad Study Material ###########
def UploadMaterial(request, cour):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    data = StudyMaterial.objects.all()
    pck = []
    cou = 0
    today_date = date.today()
    format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
    batches = []
    if cour != "0":
        cou = Master_course_data.objects.filter(id=cour).first()
        batches = Master_batch_data.objects.filter(course=cou)
    if request.method == "POST":
        b = request.POST.getlist('checks')
        name = request.POST["name"]
        file = request.FILES["file"]
        material = StudyMaterial.objects.create(course = cou, name = name, date = format_date, file = file )
        for i in b:
            bch = Master_batch_data.objects.filter(id = i).first()
            Material_Batch.objects.create(batch = bch, material = material)
        return redirect("ExamPanel:upload_material", '0')

    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'batches': batches,'packages':data,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
        }
    return render(request, "exam/uploadMaterial.html", context)



def DeleteUploadMaterila(request, mid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    StudyMaterial.objects.filter(id = mid).delete()
    return redirect(request.META.get('HTTP_REFERER'))



######## Daily Study Material ###########
def DailyPractice(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = DailyPracticeSubject.objects.all().order_by('-id')
    form = DailyPracticeSubjectForm()
    if request.method == "POST":
        form = DailyPracticeSubjectForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:daily_practice_material")
    return render(request, "Practice/DailyPractice.html", {"packages":data, "form":form,
                                                           "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditDailyMaterial(request, eid, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if what == "Edit":
        data = DailyPracticeSubject.objects.filter(id = eid).first()
        form = DailyPracticeSubjectForm(request.POST or None, request.FILES or None, instance=data)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:daily_practice_material")

        return render(request, "Practice/edit_daily_material.html", {"form":form,
                                                                     "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})
    if what == "Delete":
        data = DailyPracticeSubject.objects.filter(id=eid).first()
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))


def PracticeTopic(request, For):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if For == "All":
        data = Topic.objects.all().order_by('-id')
    if For != "All":
        practice = DailyPracticeSubject.objects.filter(id = For).first()
        data = Topic.objects.filter(package = practice).order_by('-id')
    form = TopicForm()
    if request.method == "POST":
        form = TopicForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:daily_practice_topic")
    return render(request, "Practice/PracticeTopic.html", {"packages":data, "form":form,
                                                           "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditPracticeTopic(request, eid, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if what == "Edit":
        data = Topic.objects.filter(id = eid).first()
        form = TopicForm(request.POST or None, request.FILES or None, instance=data)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:daily_practice_topic")

        return render(request, "Practice/edit_daily_practice_topic.html", {"form":form,"admin":admin_show(),
                                                                           "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})
    if what == "Delete":
        data = Topic.objects.filter(id=eid).first()
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))


def AddQuestion_In_Topic(request, q_id, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(q_id) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    sections = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        sections = Question_Section.objects.filter(course = cour)
        if sid != "0":
            section = Question_Section.objects.filter(id=sid).first()
            all_ques_tion = MultipleChoiceQuestions.objects.filter(section = section)

            topic = Topic.objects.filter(id=q_id).first()
            pre_que = Topic_Question.objects.filter(topic=topic)


            for i in all_ques_tion:
                n = 0
                for j in pre_que:
                    if i.id == j.Question.id:
                        n += 1
                if n == 0:
                    ques_tion.append(i)

            page = request.GET.get('page', 1)
            num = int(pno) * (int(page)-1)
            paginator = Paginator(ques_tion, int(pno))
            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)

            if request.method == "POST":
                qu = request.POST.getlist('que')
                sec = Topic.objects.filter(id=q_id).first()
                for i in qu:
                    que = MultipleChoiceQuestions.objects.filter(id = i).first()
                    if que:
                        Topic_Question.objects.create(topic = sec, Question = que)
                return redirect("ExamPanel:add_questions_in_topic", q_id, cid, sid, pno)

    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'sections':sections, 'sec':sid, "isec":int(sid),
        'questions': data, 'users': data, "complete_url":complete_url,
        'pno':pno, 'cid':cid, 'num':num, 's_id':q_id,"admin":admin_show(),
        "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Practice/Add_Question.html", context)



def TopicQuestionList(request, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if sid != "0":
        sec = Topic.objects.filter(id = sid).first()
        all_eq = Topic_Question.objects.filter(topic = sec)
        for i in all_eq:
            ques_tion.append(i)

        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'questions': data, 'users': data,
        'pno':pno, 'num':num, 'cid':sid, 'complete_url':complete_url,"admin":admin_show(),
        "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Practice/TopicQuestionList.html", context)



################ Students and Assign Packages
def Institute_students(request,cor,bat):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = Master_course_data.objects.all()
    students = False
    batch = False
    course = False
    if cor != '0':
        course = Master_course_data.objects.filter(id= cor).first()
        students = Front_student_course_batch_data.objects.filter(course = course)
    if bat != '0':
        batch = Master_batch_data.objects.filter(id = bat).first()
        students = Front_student_course_batch_data.objects.filter(course = course,Batch =batch )

    context = {"all_course":all_course,"students":students,"course":course,"batch":batch
               ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()}
    return render(request,'Practice/institute_filter_students.html',context)



def Institute_view_student(request,option,sid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    student_data = Front_student_data.objects.filter(id = sid).first()
    print(student_data)
    if option == "View":
        num = []
        a = 0
        for sc in student_data.front_student_course_batch_data_set.all():
            for ftype in sc.front_student_fee_type_data_set.all():
                a = a + int(ftype.fee)
            num.append(a)
            a = 0
        return render(request,'Practice/institute_student_details.html',{"student_data":student_data,"total":num,
                                                                         "sid":sid,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def Assign_test_package_Students(request, stu_id, cour, test, decide):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = Sub_course.objects.all()
    exam = []
    tst = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        exam = ExamPackage.objects.filter(course = cou, status = "Publish")
        if test != "0":
            tst = ExamPackage.objects.filter(id = test).first()
            if request.method == "POST":
                today_date = date.today()
                format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
                total = int(request.POST["total_fee"])
                paid = int(request.POST["paid"])
                remain = total-paid

                exam_package = ExamPackage.objects.filter(id = test).first()
                '''Create Exam Package'''
                e_p = StudentExamPackage.objects.create(course = exam_package.course, name = exam_package.name,
                            fee = exam_package.fee, o_fee = exam_package.o_fee, logo = exam_package.logo, syllabus = exam_package.syllabus)

                package_sub_package = Package_Subpackage.objects.filter(package = exam_package)
                for p_s in package_sub_package:
                    exam_sub_package = ExamSubPackage.objects.filter(id = p_s.sub.id).first()
                    e_s_p = StudentExamSubPackage.objects.create(course = exam_sub_package.course,
                                                        name = exam_sub_package.name)

                    s_p_s = StudentPackage_Subpackage.objects.create(package = e_p, sub = e_s_p)

                    sub_package_main = SubPackage_MainExam.objects.filter(sub = exam_sub_package)
                    for s_m_e in sub_package_main:
                        main = MainExam.objects.filter(id = s_m_e.main.id).first()

                        s_main = StudentMainExam.objects.create(course = main.course, main = main, name = main.name,
                                  start_date = main.start_date, end_date = main.end_date, duration = main.duration,
                                                instruction = main.instruction, syllabus = main.syllabus)
                        if decide == "Institute":
                            stu = Front_student_data.objects.filter(id=stu_id).first()
                            s_main.IStudent = stu
                            s_main.save()
                        if decide == "External":
                            stu = Add_New_Student.objects.filter(id=stu_id).first()
                            s_main.EStudent = stu
                            s_main.save()

                        s_p_m = StudentSubPackage_MainExam.objects.create(sub = e_s_p, main = s_main)

                        main_exam_section = MainExamSection.objects.filter(exam = main)
                        for section in main_exam_section:
                            s_e_s = StudentMainExamSection.objects.create(exam = s_main, name = section.name)

                            questions = ExamQuestion.objects.filter(MainSection = section)
                            for q in questions:
                                que = MultipleChoiceQuestions.objects.filter(id = q.Question.id).first()
                                sque = CreateNewQuestion(que)
                                StudentExamQuestion.objects.create(MainSection = s_e_s, Question = sque)

                number = InvoiceNumber.objects.filter().first()
                i_n = int(number.num) + 1
                number.num = i_n
                number.save()
                i_n =  number.start_char + str(i_n)
                if decide == "Institute":
                    stu = Front_student_data.objects.filter(id=stu_id).first()
                    assign = Assign_Test_Series_Institute_Student.objects.create(student=stu, package=e_p, real = exam_package,
                                                                        total_fee=e_p.fee, paid_fee=paid,
                                                                        remain_fee=remain, date=format_date)

                    ExamInvoiceGenerate.objects.create(invoice = i_n, exam = assign, total_fee = e_p.fee,
                                                       paid_fee=paid, remain_fee=remain, date=format_date, mode = "Offline")
                    Send_msg_for_new_test_series_asign(assign)
                    return redirect("ExamPanel:institute_students_view", "View", stu_id)
                if decide == "External":
                    stu = Add_New_Student.objects.filter(id=stu_id).first()
                    assign = Assign_Test_Series_External_Student.objects.create(student=stu, package=e_p, real = exam_package,
                                                                        total_fee=e_p.fee, paid_fee=paid,
                                                                        remain_fee=remain,
                                                                        date=format_date)
                    ExamInvoiceGenerate.objects.create(invoice=i_n, o_exam=assign, total_fee=e_p.fee,
                                                       paid_fee=paid, remain_fee=remain, date=format_date,
                                                       mode="Offline")
                    Send_msg_for_new_test_series_asign(assign)

                    return redirect("ExamPanel:other_students_view", "View", stu_id)

                return redirect("ExamPanel:institute_students_view", "View", stu_id)

    context = {
        "all_course": all_course,
        's_id':stu_id, 'cour':cour, 'icour':int(cour),
        'test':test, 'itest':int(test),
        'exams':exam, 'pack':tst, 'decide':decide,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/asign_test_series.html", context)




def Assign_quiz_package_Students(request, stu_id, cour, quiz, decide):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = Sub_course.objects.all()
    exam = []
    tst = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        exam = QuizPackage.objects.filter(course = cou, status = "Publish")
        if quiz != "0":
            tst = QuizPackage.objects.filter(id = quiz).first()
            if request.method == "POST":
                today_date = date.today()
                format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
                total = int(request.POST["total_fee"])
                paid = int(request.POST["paid"])
                remain = total-paid
                print(remain)

                quiz_package = QuizPackage.objects.filter(id = quiz).first()
                quiz_quizpack = Quiz_QuizPackage.objects.filter(package=quiz_package)
                q_p = StudentQuizPackage.objects.create(course = quiz_package.course, name = quiz_package.name,
                            logo = quiz_package.logo, fee = quiz_package.fee, o_fee = quiz_package.o_fee, syllabus = quiz_package.syllabus)


                if quiz_quizpack:
                    for qu in quiz_quizpack:
                        quiz = Quiz.objects.filter(id = qu.quiz.id).first()

                        s_q = StudentQuiz.objects.create(course = quiz.course, main = quiz, name = quiz.name, duration = quiz.duration,
                                    start_date = quiz.start_date, end_date = quiz.end_date)

                        if decide == "Institute":
                            stu = Front_student_data.objects.filter(id=stu_id).first()
                            s_q.IStudent = stu
                            s_q.save()

                        if decide == "External":
                            stu = Add_New_Student.objects.filter(id=stu_id).first()
                            s_q.EStudent = stu
                            s_q.save()

                        s_qq_p = StudentQuiz_QuizPackage.objects.create(package = q_p, quiz = s_q)

                        questions = Quiz_Questions.objects.filter(quiz = quiz)
                        if questions:
                            for i in questions:
                                que = MultipleChoiceQuestions.objects.filter(id = i.Question.id).first()
                                sque = CreateNewQuestion(que)
                                StudentQuiz_Questions.objects.create(quiz = s_q, Question = sque)

                number = InvoiceNumber.objects.filter().first()
                i_n = int(number.num) + 1
                number.num = i_n
                number.save()
                i_n = number.start_char + str(i_n)

                if decide == "Institute":
                    stu = Front_student_data.objects.filter(id=stu_id).first()
                    assign = Assign_Quiz_Institute_Student.objects.create(student=stu, package=q_p, real = quiz_package,
                                                                        total_fee=q_p.fee, paid_fee=paid, remain_fee=remain,
                                                                        date=format_date)
                    ExamInvoiceGenerate.objects.create(invoice=i_n, quiz=assign, total_fee=q_p.fee,
                                                       paid_fee=paid, remain_fee=remain, date=format_date,
                                                       mode="Offline")
                    Send_msg_for_new_test_series_asign(assign)
                    return redirect("ExamPanel:institute_students_view", "View", stu_id)

                if decide == "External":
                    stu = Add_New_Student.objects.filter(id=stu_id).first()
                    assign = Assign_Quiz_External_Student.objects.create(student=stu, package=q_p, real = quiz_package,
                                                                 total_fee=q_p.fee, paid_fee=paid, remain_fee=remain,
                                                                 date=format_date)
                    ExamInvoiceGenerate.objects.create(invoice=i_n, o_quiz=assign, total_fee=q_p.fee,
                                                       paid_fee=paid, remain_fee=remain, date=format_date,
                                                       mode="Offline")
                    Send_msg_for_new_test_series_asign(assign)
                    return redirect("ExamPanel:other_students_view", "View", stu_id)

    context = {
        "all_course": all_course,
        's_id':stu_id, 'cour':cour, 'icour':int(cour),
        'test':quiz, 'itest':int(quiz),
        'exams':exam, 'pack':tst, 'decide':decide, "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/asign_quiz.html", context)




def Assign_Study_Material_Students(request, stu_id, cour, study, decide ):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = Sub_course.objects.all()
    exam = []
    tst = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        exam = DailyPracticeSubject.objects.filter(course = cou)
        if study != "0":
            today_date = date.today()
            format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
            tst = DailyPracticeSubject.objects.filter(id = study).first()

            if decide == "Institute":
                stu = Front_student_data.objects.filter(id=stu_id).first()
                Assign_Study_Institute_Student.objects.create(student=stu, package=tst,
                                                              date=format_date)

                return redirect("ExamPanel:institute_students_view", "View", stu_id)
            if decide == "External":
                stu = Add_New_Student.objects.filter(id=stu_id).first()
                Assign_Study_External_Student.objects.create(student=stu, package=tst,
                                                              date=format_date)
                return redirect("ExamPanel:other_students_view", "View", stu_id)

    context = {
        "all_course": all_course,
        's_id':stu_id, 'cour':cour, 'icour':int(cour),
        'exams':exam, 'pack':tst, 'decide':decide,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/asign_study.html", context)



def Cancel_all(request, what, decide):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if request.method == "POST":
        if what == "Test":
            val = request.POST["test"]
            if decide == "Institute":
                data = Assign_Test_Series_Institute_Student.objects.filter(id = val)
                if request.user.is_staff:
                    data.delete()
            if decide == "External":
                data = Assign_Test_Series_External_Student.objects.filter(id=val)
                if request.user.is_staff:
                    data.delete()
        if what == "Quiz":
            val = request.POST["quiz"]
            if decide == "Institute":
                data = Assign_Quiz_Institute_Student.objects.filter(id = val)
                if request.user.is_staff:
                    data.delete()
            if decide == "External":
                data = Assign_Quiz_External_Student.objects.filter(id=val)
                if request.user.is_staff:
                    data.delete()
        if what == "Study":
            val = request.POST["study"]
            if decide == "Institute":
                data = Assign_Study_Institute_Student.objects.filter(id = val)
                if request.user.is_staff:
                    data.delete()
            if decide == "External":
                data = Assign_Study_External_Student.objects.filter(id=val)
                if request.user.is_staff:
                    data.delete()
        return redirect(request.META.get('HTTP_REFERER'))




def Pay_Test_Quiz_Fee(request, pid, what, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if request.method == "POST":
        number = InvoiceNumber.objects.filter().first()
        i_n = int(number.num) + 1
        number.num = i_n
        number.save()

        today_date = date.today()
        format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')

        if what == "Test":
            if option == "Institute":
                data = Assign_Test_Series_Institute_Student.objects.filter(id = pid).first()
                amount = int(request.POST['amount'])
                remain = int(data.remain_fee)-amount
                ExamInvoiceGenerate.objects.create(invoice=i_n, exam=data, total_fee=data.remain_fee,
                                                   paid_fee=amount, remain_fee=remain, date=format_date,
                                                   mode="Offline")
                data.paid_fee = int(data.paid_fee) + amount
                data.remain_fee = remain
                data.save()

            if option == "External":
                data = Assign_Test_Series_External_Student.objects.filter(id=pid).first()
                amount = int(request.POST['amount'])
                remain = int(data.remain_fee) - amount
                ExamInvoiceGenerate.objects.create(invoice=i_n, o_exam=data, total_fee=data.remain_fee,
                                                   paid_fee=amount, remain_fee=remain, date=format_date,
                                                   mode="Offline")
                data.paid_fee = int(data.paid_fee) + amount
                data.remain_fee = remain
                data.save()
        if what == "Quiz":
            if option == "Institute":
                data = Assign_Quiz_Institute_Student.objects.filter(id=pid).first()
                amount = int(request.POST['amount'])
                remain = int(data.remain_fee) - amount
                ExamInvoiceGenerate.objects.create(invoice=i_n, quiz=data, total_fee=data.remain_fee,
                                                   paid_fee=amount, remain_fee=remain, date=format_date,
                                                   mode="Offline")
                data.paid_fee = int(data.paid_fee) + amount
                data.remain_fee = remain
                data.save()
            if option == "External":
                data = Assign_Quiz_External_Student.objects.filter(id=pid).first()
                amount = int(request.POST['amount'])
                remain = int(data.remain_fee) - amount
                ExamInvoiceGenerate.objects.create(invoice=i_n, o_quiz=data, total_fee=data.remain_fee,
                                                   paid_fee=amount, remain_fee=remain, date=format_date,
                                                   mode="Offline")
                data.paid_fee = int(data.paid_fee) + amount
                data.remain_fee = remain
                data.save()
        return redirect(request.META.get('HTTP_REFERER'))

from django.core.mail import send_mail, EmailMultiAlternatives, EmailMessage
from django.conf import settings
from django.template.loader import get_template, render_to_string
def Admission_Mail(email, institute, student, username, Pass ):
    from_email = settings.EMAIL_HOST_USER
    to_email = [email]
    d = {
        'institute':institute,
        'student':student, 'username':username, 'pass':Pass
    }
    html = get_template("add_new_student_mail.html").render(d)
    sub = institute.Title + " - New Admission"
    msg = EmailMultiAlternatives(sub, " ", from_email, to_email)
    msg.attach_alternative(html, "text/html")
    msg.send()

def AddNewStudents(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = Add_New_Student.objects.all().order_by('-id')
    form = AddNewStudentForm()
    today_date = date.today()
    format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
    if request.method == "POST":
        form = AddNewStudentForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.date = format_date
            data.save()

            uname = 0
            while True:
                name = data.name
                name = name.split()[0]
                num = random.randint(100000, 999999)
                uname = name + str(num)
                usr = User.objects.filter(username=uname).first()
                if not usr:
                    break
            Pass = data.mobile[:6]
            user = User.objects.create_user(uname, data.email, Pass)
            user.last_name = "E_Student"
            user.first_name = data.name
            user.save()
            data.usr = user
            data.save()
            Admission_Mail(data.email, Institute_data(), data.name, uname, Pass)
            return redirect("ExamPanel:add_new_students")
    return render(request, "Students/add_new_student.html", {"form":form, 'packages':data, 'today':format_date,
                                                             "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})

def Edit_Delete_Other_Students(request, sid, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = Add_New_Student.objects.filter(id = sid).first()
    if what == "Delete":
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))
    if what == "Edit":
        form = AddNewStudentForm(request.POST or None, request.FILES or None, instance=data)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:add_new_students")
        return render(request, "Students/edit_new_student.html", {"form": form, 'packages': data
                                                                  ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def Other_view_student(request,option,sid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    student_data = Add_New_Student.objects.filter(id = sid).first()
    if not student_data:
        return redirect(request.META.get('HTTP_REFERER'))
    if option == "View":
        return render(request,'Practice/other_student_details.html',{"student_data":student_data, 'sid':sid
                                                                     ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def Student_Groups(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    group = StudentGroup.objects.all()
    today_date = date.today()
    format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
    if request.method == "POST":
        name = request.POST['name']
        StudentGroup.objects.create(date = format_date, name = name)
        return redirect("ExamPanel:new_students_group")
    return render(request, "Students/student_group.html", {'packages': group,
                                                           "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def Student_Groups_Delete(request, gid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    group = StudentGroup.objects.filter(id = gid).first()
    group.delete()
    return redirect(request.META.get('HTTP_REFERER'))


def Add_Students_In_Group(request, gid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    students = []
    all_students = Add_New_Student.objects.all()
    group = StudentGroup.objects.filter(id = gid).first()
    group_students = Group_Students.objects.all()
    for s in all_students:
        n = 0
        for j in group_students:
            if s.id == j.student.id:
                n += 1

        if n == 0:
            students.append(s)
    if request.method == "POST":
        d = request.POST.getlist('checks')
        print(d)
        for i in d:
            s = Add_New_Student.objects.filter(id = i).first()
            Group_Students.objects.create(group = group, student = s)
        return redirect("ExamPanel:add_students_in_group", gid)

    return render(request, "Students/select_students.html" ,{'students':students,
                                                             "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def Group_Student_List(request, gid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    group = StudentGroup.objects.filter(id = gid).first()
    return render(request, "Students/group_students_list.html", {"group":group,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def Assign_Test_Function(students, eid, option):

    for student in students:
        today_date = date.today()
        format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')

        exam_package = ExamPackage.objects.filter(id=eid).first()
        '''Create Exam Package'''
        e_p = StudentExamPackage.objects.create(course=exam_package.course, name=exam_package.name,
                                                fee=exam_package.fee, o_fee=exam_package.o_fee,
                                                logo=exam_package.logo, syllabus=exam_package.syllabus)

        package_sub_package = Package_Subpackage.objects.filter(package=exam_package)
        for p_s in package_sub_package:
            exam_sub_package = ExamSubPackage.objects.filter(id=p_s.sub.id).first()
            e_s_p = StudentExamSubPackage.objects.create(course=exam_sub_package.course,
                                                         name=exam_sub_package.name)

            s_p_s = StudentPackage_Subpackage.objects.create(package=e_p, sub=e_s_p)

            sub_package_main = SubPackage_MainExam.objects.filter(sub=exam_sub_package)
            for s_m_e in sub_package_main:
                main = MainExam.objects.filter(id=s_m_e.main.id).first()

                s_main = StudentMainExam.objects.create(course=main.course, main = main, name=main.name,
                                                        start_date=main.start_date,
                                                        end_date=main.end_date, duration=main.duration,
                                                        instruction=main.instruction,
                                                        syllabus=main.syllabus)
                if option == "Institute":
                    stu = Front_student_data.objects.filter(id=student.id).first()
                    s_main.IStudent = stu
                    s_main.save()
                if option == "External":
                    stu = Add_New_Student.objects.filter(id=student.id).first()
                    s_main.EStudent = stu
                    s_main.save()

                s_p_m = StudentSubPackage_MainExam.objects.create(sub=e_s_p, main=s_main)

                main_exam_section = MainExamSection.objects.filter(exam=main)
                for section in main_exam_section:
                    s_e_s = StudentMainExamSection.objects.create(exam=s_main, name=section.name)

                    questions = ExamQuestion.objects.filter(MainSection=section)
                    for q in questions:
                        que = MultipleChoiceQuestions.objects.filter(id=q.Question.id).first()
                        sque = CreateNewQuestion(que)
                        StudentExamQuestion.objects.create(MainSection=s_e_s, Question=sque)

        number = InvoiceNumber.objects.filter().first()
        i_n = int(number.num) + 1
        number.num = i_n
        number.save()
        i_no = number.start_char + str(i_n)
        if option == "Institute":
            stu = Front_student_data.objects.filter(id=student.id).first()
            assign = Assign_Test_Series_Institute_Student.objects.create(student=stu, package=e_p, real = exam_package,
                                                                         total_fee=e_p.fee,
                                                                         paid_fee=e_p.fee,
                                                                         remain_fee=0,
                                                                         date=format_date)

            pid = ExamInvoiceGenerate.objects.create(invoice=i_no, exam=assign, total_fee=e_p.fee,
                                               paid_fee=e_p.fee, remain_fee=0, date=format_date,
                                               mode="Offline")
            Send_msg_for_new_test_series_asign(assign)
            return redirect('ExamPanel:exam_generate_invoice',pid.id)

        if option == "External":
            stu = Add_New_Student.objects.filter(id=student.id).first()
            assign = Assign_Test_Series_External_Student.objects.create(student=stu, package=e_p, real = exam_package,
                                                                        total_fee=e_p.fee,
                                                                        paid_fee=e_p.fee,
                                                                        remain_fee=0,
                                                                        date=format_date)
            pid = ExamInvoiceGenerate.objects.create(invoice=i_no, o_exam=assign, total_fee=e_p.fee,
                                               paid_fee=e_p.fee, remain_fee=0, date=format_date,
                                               mode="Offline")
            Send_msg_for_new_test_series_asign(assign)
            return redirect('ExamPanel:exam_generate_invoice', pid.id)


def Assign_Test_Series_to_Groups(request, cid, bid, eid, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    batches = []
    exams = []
    students = []
    Groups = []
    if option == "Institute":
        if cid != '0':
            cou = Sub_course.objects.filter(id = cid).first()
            batches = Master_batch_data.objects.filter(course = cou)
            if bid != '0':
                batch = Master_batch_data.objects.filter(id = bid).first()
                students = Front_student_course_batch_data.objects.filter(course=cou, Batch=batch)
                exams = ExamPackage.objects.filter(status = "Publish")
            if eid != '0':
                if request.method == "POST":
                    add_students = []
                    d = request.POST.getlist("checks")
                    if len(d)>0:
                        for i in d:
                            st = Front_student_data.objects.filter(id = i).first()
                            add_students.append(st)
                        Assign_Test_Function(add_students, eid, option)
                        return redirect("ExamPanel:exam_home")
                    else:
                        return redirect(request.META.get('HTTP_REFERER'))

    if option == "External":
        Groups = StudentGroup.objects.all()
        if cid != '0':
            cou = Sub_course.objects.filter(id=cid).first()
            exams = ExamPackage.objects.filter(course=cou, status = "Publish")

        if bid != '0':
            group = StudentGroup.objects.filter(id = bid).first()
            student = Group_Students.objects.filter(group = group)
            students = []
            for i in student:
                stu = Add_New_Student.objects.filter(id = i.student.id).first()
                students.append(stu)
            if request.method == "POST":
                add_students = []
                d = request.POST.getlist("checks")
                if len(d) > 0:
                    for i in d:
                        print(d)
                        st = Add_New_Student.objects.filter(id=i).first()
                        add_students.append(st)
                    Assign_Test_Function(add_students, eid, option)
                    return redirect("ExamPanel:exam_home")
                else:
                    return redirect(request.META.get('HTTP_REFERER'))


    context = {
        'all_course':course, 'batches':batches, 'exams':exams,
        'cid':cid, 'bid':bid, 'eid':eid, "option":option, 'icid':int(cid), 'ibid':int(bid),
        'ieid':int(eid), 'students':students,
        'groups':Groups,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/assign_test_to_groups.html", context)




def Assign_Quiz_Function(students, quiz, option):
    for student in students:
        today_date = date.today()
        format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')

        quiz_package = QuizPackage.objects.filter(id=quiz).first()
        quiz_quizpack = Quiz_QuizPackage.objects.filter(package=quiz_package)
        q_p = StudentQuizPackage.objects.create(course=quiz_package.course, name=quiz_package.name,
                                                logo=quiz_package.logo, fee=quiz_package.fee, o_fee=quiz_package.o_fee,
                                                syllabus=quiz_package.syllabus)

        if quiz_quizpack:
            for qu in quiz_quizpack:
                print(qu.quiz.name)
                quiz = Quiz.objects.filter(id=qu.quiz.id).first()

                s_q = StudentQuiz.objects.create(course=quiz.course, main = quiz, name=quiz.name, duration=quiz.duration,
                                                 start_date=quiz.start_date, end_date=quiz.end_date)
                if option == "Institute":
                    stu = Front_student_data.objects.filter(id=student.id).first()
                    s_q.IStudent = stu
                    s_q.save()

                if option == "External":
                    stu = Add_New_Student.objects.filter(id=student.id).first()
                    s_q.EStudent = stu
                    s_q.save()

                s_qq_p = StudentQuiz_QuizPackage.objects.create(package=q_p, quiz=s_q)

                questions = Quiz_Questions.objects.filter(quiz=quiz)
                if questions:
                    for i in questions:
                        que = MultipleChoiceQuestions.objects.filter(id=i.Question.id).first()
                        sque = CreateNewQuestion(que)
                        StudentQuiz_Questions.objects.create(quiz=s_q, Question=sque)

        number = InvoiceNumber.objects.filter().first()
        i_n = int(number.num) + 1
        number.num = i_n
        number.save()
        i_no = number.start_char + str(i_n)

        if option == "Institute":
            stu = Front_student_data.objects.filter(id=student.id).first()
            assign = Assign_Quiz_Institute_Student.objects.create(student=stu, package=q_p, real = quiz_package,
                                                                  total_fee=q_p.fee, paid_fee=q_p.fee, remain_fee=0,
                                                                  date=format_date)
            pid = ExamInvoiceGenerate.objects.create(invoice=i_no, quiz=assign, total_fee=q_p.fee,
                                               paid_fee=q_p.fee, remain_fee=0, date=format_date,
                                               mode="Offline")
            Send_msg_for_new_test_series_asign(assign)

            return redirect('ExamPanel:exam_generate_invoice', pid.id)


        if option == "External":
            stu = Add_New_Student.objects.filter(id=student.id).first()
            assign = Assign_Quiz_External_Student.objects.create(student=stu, package=q_p, real = quiz_package,
                                                                 total_fee=q_p.fee, paid_fee=q_p.fee, remain_fee=0,
                                                                 date=format_date)
            pid = ExamInvoiceGenerate.objects.create(invoice=i_no, o_quiz=assign, total_fee=q_p.fee,
                                               paid_fee=q_p.fee, remain_fee=0, date=format_date,
                                               mode="Offline")
            Send_msg_for_new_test_series_asign(assign)
            return redirect('ExamPanel:exam_generate_invoice', pid.id)



def Assign_Quizes_to_Groups(request, cid, bid, eid, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    batches = []
    exams = []
    students = []
    Groups = []
    if option == "Institute":
        if cid != '0':
            cou = Sub_course.objects.filter(id = cid).first()
            batches = Master_batch_data.objects.filter(course = cou)
            if bid != '0':
                batch = Master_batch_data.objects.filter(id = bid).first()
                students = Front_student_course_batch_data.objects.filter(course=cou, Batch=batch)
                exams = QuizPackage.objects.filter(course=cou, status = "Publish")
            if eid != '0':
                if request.method == "POST":
                    add_students = []
                    d = request.POST.getlist("checks")
                    if len(d)>0:
                        for i in d:
                            st = Front_student_data.objects.filter(id = i).first()
                            add_students.append(st)
                        Assign_Quiz_Function(add_students, eid, option)
                        return redirect("ExamPanel:exam_home")
                    else:
                        return redirect(request.META.get('HTTP_REFERER'))

    if option == "External":
        Groups = StudentGroup.objects.all()
        if cid != '0':
            cou = Sub_course.objects.filter(id=cid).first()
            exams = QuizPackage.objects.filter(course=cou, status = "Publish" )

        if bid != '0':
            group = StudentGroup.objects.filter(id = bid).first()
            student = Group_Students.objects.filter(group = group)
            students = []
            for i in student:
                stu = Add_New_Student.objects.filter(id = i.student.id).first()
                students.append(stu)
            if request.method == "POST":
                add_students = []
                d = request.POST.getlist("checks")
                if len(d) > 0:
                    for i in d:
                        print(d)
                        st = Add_New_Student.objects.filter(id=i).first()
                        add_students.append(st)
                        Assign_Quiz_Function(add_students, eid, option)
                    return redirect("ExamPanel:exam_home")
                else:
                    return redirect(request.META.get('HTTP_REFERER'))


    context = {
        'all_course':course, 'batches':batches, 'exams':exams,
        'cid':cid, 'bid':bid, 'eid':eid, "option":option, 'icid':int(cid), 'ibid':int(bid),
        'ieid':int(eid), 'students':students,
        'groups':Groups,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/assign_quiz_to_groups.html", context)



def Assign_Study_Material_Groups(request, cid, bid, eid, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    batches = []
    exams = []
    students = []
    Groups = []
    today_date = date.today()
    format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
    if option == "Institute":
        if cid != '0':
            cou = Sub_course.objects.filter(id = cid).first()
            batches = Master_batch_data.objects.filter(course = cou)
            if bid != '0':
                batch = Master_batch_data.objects.filter(id = bid).first()
                students = Front_student_course_batch_data.objects.filter(course=cou, Batch=batch)
                exams = DailyPracticeSubject.objects.filter(course=cou)
            if eid != '0':
                if request.method == "POST":
                    d = request.POST.getlist("checks")
                    tst = DailyPracticeSubject.objects.filter(id=eid).first()
                    print(tst)
                    if len(d)>0:
                        for i in d:
                            stu = Front_student_data.objects.filter(id=i).first()
                            Assign_Study_Institute_Student.objects.create(student=stu, package=tst,
                                                                          date=format_date)
                            print(stu.name, stu.id)
                        return redirect("ExamPanel:exam_home")
                    else:
                        return redirect(request.META.get('HTTP_REFERER'))

    if option == "External":
        Groups = StudentGroup.objects.all()
        if cid != '0':
            cou = Sub_course.objects.filter(id=cid).first()
            exams = DailyPracticeSubject.objects.filter(course=cou)

        if bid != '0':
            group = StudentGroup.objects.filter(id = bid).first()
            student = Group_Students.objects.filter(group = group)
            students = []
            for i in student:
                stu = Add_New_Student.objects.filter(id = i.student.id).first()
                students.append(stu)
            if request.method == "POST":
                tst = DailyPracticeSubject.objects.filter(id=eid).first()
                d = request.POST.getlist("checks")
                if len(d) > 0:
                    for i in d:
                        stu = Add_New_Student.objects.filter(id=i).first()
                        Assign_Study_External_Student.objects.create(student=stu, package=tst,
                                                                     date=format_date)
                    return redirect("ExamPanel:exam_home")
                else:
                    return redirect(request.META.get('HTTP_REFERER'))


    context = {
        'all_course':course, 'batches':batches, 'exams':exams,
        'cid':cid, 'bid':bid, 'eid':eid, "option":option, 'icid':int(cid), 'ibid':int(bid),
        'ieid':int(eid), 'students':students,
        'groups':Groups,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/assign_study_to_groups.html", context)



def Assigned_Students(request, eid, what, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    students = []
    option = "Institute"
    num = 0
    if option == "Institute":
        students = []
        if what == "Exam":
            exam = ExamPackage.objects.filter(id = eid).first()
            stu = Assign_Test_Series_Institute_Student.objects.filter(real = exam)
            for i in stu:
                st = Front_student_data.objects.filter(id = i.student.id).first()
                students.append(st)

        if what == "Quiz":
            exam = QuizPackage.objects.filter(id=eid).first()
            stu = Assign_Quiz_Institute_Student.objects.filter(real=exam)
            for i in stu:
                st = Front_student_data.objects.filter(id=i.student.id).first()
                students.append(st)
        if what == "Study":
            exam = DailyPracticeSubject.objects.filter(id=eid).first()
            stu = Assign_Study_Institute_Student.objects.filter(package=exam)
            for i in stu:
                st = Front_student_data.objects.filter(id=i.student.id).first()
                students.append(st)


    if option == "External":
        students = []
        if what == "Exam":
            exam = ExamPackage.objects.filter(id=eid).first()
            stu = Assign_Test_Series_External_Student.objects.filter(real=exam)
            for i in stu:
                st = Add_New_Student.objects.filter(id=i.student.id).first()
                students.append(st)
        if what == "Quiz":
            exam = QuizPackage.objects.filter(id=eid).first()
            stu = Assign_Quiz_External_Student.objects.filter(real=exam)
            for i in stu:
                st = Front_student_data.objects.filter(id=i.student.id).first()
                students.append(st)
        if what == "Study":
            exam = DailyPracticeSubject.objects.filter(id=eid).first()
            stu = Assign_Study_External_Student.objects.filter(package=exam)
            for i in stu:
                st = Add_New_Student.objects.filter(id=i.student.id).first()
                students.append(st)

    context = {
        'option':option, 'eid':eid, 'students':students, 'what':what, 'num':len(students)
        ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "Students/assigned_students.html", context)






from django.http import HttpRequest
def Home(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    courses = Sub_course.objects.all()
    exams = ExamPackage.objects.all()
    quizs = QuizPackage.objects.all()
    practice = Quiz.objects.all()
    all_students = Front_student_data.objects.all()
    other_students = Add_New_Student.objects.all().order_by('-id')
    questions = MultipleChoiceQuestions.objects.all()
    last_five_new_students = other_students[:5]
    today_date = date.today()
    format_date = datetime.strptime(str(today_date), "%Y-%m-%d").strftime('%d/%m/%Y')
    assign1 = Assign_Test_Series_Institute_Student.objects.filter(date = format_date)
    assign2 = Assign_Test_Series_External_Student.objects.filter(date = format_date)
    test_income = 0
    for i in assign1:
        test_income += int(i.paid_fee)
    for i in assign2:
        test_income += int(i.paid_fee)
    total_test = len(assign1) + len(assign2)
    assign1 = Assign_Quiz_Institute_Student.objects.filter(date=format_date)
    assign2 = Assign_Quiz_External_Student.objects.filter(date=format_date)
    total_quiz = len(assign1) + len(assign2)
    quiz_income = 0
    for i in assign1:
        quiz_income += int(i.paid_fee)
    for i in assign2:
        quiz_income += int(i.paid_fee)
    d = Front_student_course_batch_data.objects.order_by('-id')
    data = d[0:5]




    context = {
        'all_course':courses, 'exams':exams, 'quizs':quizs, 'practices':practice,
        'all_students':all_students, 'other_students':other_students, 'questions':questions,
        'data':last_five_new_students, 'todays_test':total_test, 'todays_quiz':total_quiz,
        'quiz_income':quiz_income, 'test_income':test_income, 'data':data,

        "ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/exam_deshboard.html", context)


def CreateExamPackage(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamPackage.objects.all().order_by('-id')
    form = CreateExamPackageForm()
    if request.method == "POST":
        form = CreateExamPackageForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            return redirect("ExamPanel:create_exam_package")
    return render(request, "exam/create_exam_package.html", {"form":form, 'packages':data
                                                             ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def EditExamPackage(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamPackage.objects.filter(id = eid).first()
    form = CreateExamPackageForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        return redirect("ExamPanel:create_exam_package")

    return render(request, "exam/edit_exam_package.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def AddExamSubPackage(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    form = CreateExamSubPackageForm()
    pack = ExamPackage.objects.filter(id = eid).first()
    sub = Package_Subpackage.objects.filter(package = pack).order_by('-id')
    courses = Sub_course.objects.all()
    if request.method == "POST":
        d = request.POST.getlist('checks')
        for i in d:
            n = 0
            for j in sub:
                if i == str(j.sub.id):
                    n = n+1
            if n == 0:
                s = ExamSubPackage.objects.filter(id = i).first()
                Package_Subpackage.objects.create(package=pack, sub = s)
        return redirect("ExamPanel:add_packages", eid)

    context = {
        'subpackages':sub, 'form1':form,
        'courses':courses,"ins_data":Institute_data(request.user), "adminprofile":Admin_data(), 'pack':pack
    }
    return render(request, "exam/sub_packages_list.html", context)



def RemoveSubPackage(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = Package_Subpackage.objects.filter(id = eid).first()
    d.delete()
    return redirect(request.META.get('HTTP_REFERER'))




def DeleteExamPackage(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamPackage.objects.filter(id = eid).first()
    if request.user.is_staff:
        data.delete()
    return redirect("ExamPanel:create_exam_package")



def CreateExamSubPackage(request, cour, pack, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamSubPackage.objects.all().order_by('-id')
    form = CreateExamSubPackageForm()
    course = Sub_course.objects.all()
    pck = []
    cou = 0
    if cour != "0":
        cou = Sub_course.objects.filter(id=cour).first()
        pck = ExamPackage.objects.filter(course=cou)
    if request.method == "POST":
        if pack == '0':
            return redirect("ExamPanel:create_exam_subpackage", cour, '0')
        form = CreateExamSubPackageForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            data.course = cou
            data.save()
            pac = ExamPackage.objects.filter(id = pack).first()
            Package_Subpackage.objects.create(package = pac, sub = data)
            if what == "New":
                return redirect("ExamPanel:create_exam_subpackage", '0', '0', "New")
            return redirect(request.META.get('HTTP_REFERER'))
    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'package': pck, 'ipack': int(pack), 'pack': pack, 'what':what,
        'exams': data, 'form': form, 'subpackages':data,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
        }
    return render(request, "exam/create_exam_subpackage.html", context)



def ExamSubPackageDetails(request, eid, cour, pack):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    sub = ExamSubPackage.objects.filter(id = eid).first()
    data = Package_Subpackage.objects.filter(sub = sub)
    course = Sub_course.objects.all()
    pck = []
    Packs = []
    if cour != "0":
        cou = Sub_course.objects.filter(id=cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        for i in pck:
            n = 0
            for j in data:
                if i == j.package:
                    n += 1
            if n == 0:
                Packs.append(i)
    if request.method == "POST":
        if cour == "0" or pack == "0":
            return redirect("ExamPanel:subpackage_details", 'eid', cour, pack)
        package = ExamPackage.objects.filter(id = pack).first()
        Package_Subpackage.objects.create(package = package, sub = sub)
        return redirect("ExamPanel:subpackage_details", eid, '0', '0')

    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'package': Packs, 'ipack': int(pack), 'pack': pack,
        'exams': data, 'packages': data, 'eid':eid,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
                }
    return render(request, "exam/exam_subpackage_details.html", context)



def EditExamSubPackage(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamSubPackage.objects.filter(id = eid).first()
    form = CreateExamSubPackageForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        return redirect("ExamPanel:create_exam_subpackage", '0', '0', 'New')

    return render(request, "exam/edit_exam_subpackage.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def DeleteExamSubPackage(request, eid):
    data = ExamSubPackage.objects.filter(id = eid).first()
    if request.user.is_staff:
        data.delete()
    return redirect(request.META.get('HTTP_REFERER'))



def Sub_Package_Exam_List(request, sid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    packages = []
    form = CreateMainExamForm()
    sub = ExamSubPackage.objects.filter(id = sid).first()
    sp = SubPackage_MainExam.objects.filter(sub = sub)
    for e in sp:
        exam = MainExam.objects.filter(id = e.main.id).first()
        packages.append(exam)
    print(packages)

    return render(request, "exam/subpackage_exam_list.html", {"exams":packages, 'sub':sub, 'form':form, "ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def CreateMainExam(request, cour, pack, subpack, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = MainExam.objects.all().order_by('-id')
    course = Sub_course.objects.all()
    d = ExamSubPackage.objects.filter(id=subpack).first()
    pck = []
    sub = []
    cou = 0
    form = CreateMainExamForm()
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        if pack != '0':
            pc = ExamPackage.objects.filter(id = pack).first()
            sub = Package_Subpackage.objects.filter(package = pc)
    if request.method == "POST":
        form = CreateMainExamForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.course = cou
            start = data.start_date
            end = data.end_date

            x = start.split('/')
            if len(x[0]) < 2:
                y = x.pop(0)
                y = '0' + y
                x.insert(0, y)
            if len(x[1]) < 2:
                y = x.pop(1)
                y = '0' + y
                x.insert(1, y)
            x = '/'.join(x)
            start = x
            x = end.split('/')
            if len(x[0]) < 2:
                y = x.pop(0)
                y = '0' + y
                x.insert(0, y)
            if len(x[1]) < 2:
                y = x.pop(1)
                y = '0' + y
                x.insert(1, y)
            x = '/'.join(x)
            end = x
            data.start_date = start
            data.end_date = end
            data.save()
            s = ExamSubPackage.objects.filter(id = subpack).first()
            SubPackage_MainExam.objects.create(sub = s, main = data)
            if what == "New":
                return redirect("ExamPanel:create_main_exam", '0', '0', '0', 'New')
            return redirect(request.META.get('HTTP_REFERER'))
    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'package': pck, 'ipack': int(pack), 'pack': pack,
        'subpackage': sub, 'isubpack': int(subpack), 'subpack': subpack,
        'exams': data, 'form':form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data(), 'what':what
    }
    return render(request, "exam/create_main_exam.html", context)



def MainExamDetails(request, eid, cour, pack, subpack):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    sub = MainExam.objects.filter(id = eid).first()
    data = SubPackage_MainExam.objects.filter(main = sub)
    course = Sub_course.objects.all()
    pck = []
    Packs = []
    if cour != "0":
        cou = Sub_course.objects.filter(id=cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        if pack != "0":
            pkg = ExamPackage.objects.filter(id = pack).first()
            subpackage = Package_Subpackage.objects.filter(package = pkg)
            for i in subpackage:
                n = 0
                for j in data:
                    if i.sub == j.sub:
                        n += 1
                if n == 0:
                    Packs.append(i)
    if request.method == "POST":
        if cour == "0" or pack == "0" or subpack == "0":
            return redirect("ExamPanel:main_exam_details", 'eid', cour, pack, subpack)
        subpackage = ExamSubPackage.objects.filter(id = subpack).first()
        SubPackage_MainExam.objects.create(sub = subpackage, main = sub)
        return redirect("ExamPanel:main_exam_details", eid, '0', '0', '0')

    context = {
        'course': course, "icour": int(cour), "cour": cour,
        'package': pck, 'ipack': int(pack), 'pack': pack,
        'subpackage':Packs, 'isubpack':int(subpack), 'subpack':subpack,
        'exams': data, 'packages': data, 'eid':eid,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
                }
    return render(request, "exam/main_exam_details.html", context)



def RemoveMainExam(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = SubPackage_MainExam.objects.filter(id = eid).first()
    d.delete()
    return redirect(request.META.get('HTTP_REFERER'))


def EditMainExam(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = MainExam.objects.filter(id = eid).first()
    form = CreateMainExamForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        start = data.start_date
        end = data.end_date

        x = start.split('/')
        if len(x[0]) < 2:
            y = x.pop(0)
            y = '0' + y
            x.insert(0, y)
        if len(x[1]) < 2:
            y = x.pop(1)
            y = '0' + y
            x.insert(1, y)
        x = '/'.join(x)
        start = x
        x = end.split('/')
        if len(x[0]) < 2:
            y = x.pop(0)
            y = '0' + y
            x.insert(0, y)
        if len(x[1]) < 2:
            y = x.pop(1)
            y = '0' + y
            x.insert(1, y)
        x = '/'.join(x)
        end = x
        data.start_date = start
        data.end_date = end
        data.save()
        return redirect("ExamPanel:create_main_exam", '0', '0', '0', 'New')
    return render(request, "exam/edit_main_exam.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditStudentMainExam(request, eid, sid, what, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = StudentMainExam.objects.filter(id = eid).first()
    if what == "Delete":
        if request.user.is_staff:
            data.delete()
        return redirect(request.META.get('HTTP_REFERER'))
    if what == "Edit":
        form = CreateMainExamForm(request.POST or None, request.FILES or None, instance=data)
        if form.is_valid():
            data = form.save(commit=False)
            start = data.start_date
            end = data.end_date

            x = start.split('/')
            if len(x[0]) < 2:
                y = x.pop(0)
                y = '0' + y
                x.insert(0, y)
            if len(x[1]) < 2:
                y = x.pop(1)
                y = '0' + y
                x.insert(1, y)
            x = '/'.join(x)
            start = x
            x = end.split('/')
            if len(x[0]) < 2:
                y = x.pop(0)
                y = '0' + y
                x.insert(0, y)
            if len(x[1]) < 2:
                y = x.pop(1)
                y = '0' + y
                x.insert(1, y)
            x = '/'.join(x)
            end = x
            data.start_date = start
            data.end_date = end
            data.save()

            if option == "Institute":
                return redirect("ExamPanel:institute_students_view", 'View', sid)
            if option == "External":
                return redirect("ExamPanel:other_students_view", 'View', sid)

    return render(request, "exam/edit_main_exam.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def StudentExamSections(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = StudentMainExam.objects.filter(id = eid).first()
    sections = StudentMainExamSection.objects.filter(exam = data)
    arg = {
        "sections":sections,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/student_exam_sections.html", arg )


def StudentExamSectionQuestionList(request, sid, pno, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(sid) + "__" + str(pno) + "__" + str(what)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if sid != "0":
        if what == "All":
            exam = StudentMainExam.objects.filter(id = sid).first()
            sections = StudentMainExamSection.objects.filter(exam = exam)
            for s in sections:
                all_eq = StudentExamQuestion.objects.filter(MainSection=s)
                for i in all_eq:
                    ques_tion.append(i)
        else:
            sec = StudentMainExamSection.objects.filter(id = sid).first()
            all_eq = StudentExamQuestion.objects.filter(MainSection = sec)
            for i in all_eq:
                ques_tion.append(i)

        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'questions': data, 'users': data,
        'pno':pno, 'num':num, 'cid':sid, "what":what, "complete_url":complete_url
        ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/StudentExamSectionQuestion.html", context)



def AddQuestion_In_StudentExamSection(request, s_id, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(s_id) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    sections = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        sections = Question_Section.objects.filter(course = cour)
        if sid != "0":
            section = Question_Section.objects.filter(id=sid).first()
            all_ques_tion = MultipleChoiceQuestions.objects.filter(section = section)

            sec = StudentMainExamSection.objects.filter(id=s_id).first()
            Exam = StudentMainExam.objects.filter(id = sec.exam.id).first()
            all_sec = StudentMainExamSection.objects.filter(exam = Exam)
            pre_que = StudentExamQuestion.objects.filter(MainSection=sec)

            for i in all_ques_tion:
                n = 0
                for z in all_sec:
                    pre_que = StudentExamQuestion.objects.filter(MainSection=z)
                    for j in pre_que:
                        if i.id == j.Question.id:
                            n+=1
                if n == 0:
                    ques_tion.append(i)

            page = request.GET.get('page', 1)
            num = int(pno) * (int(page)-1)
            paginator = Paginator(ques_tion, int(pno))
            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)

            if request.method == "POST":
                print("Hello")
                qu = request.POST.getlist('que')
                sec = StudentMainExamSection.objects.filter(id=s_id).first()
                for i in qu:
                    que = MultipleChoiceQuestions.objects.filter(id = i).first()
                    if que:
                        StudentExamQuestion.objects.create(MainSection = sec, Question = que)
                return redirect("ExamPanel:add_questions_in_student_exam_section", s_id, cid, sid, pno)


    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'sections':sections, 'sec':sid, "isec":int(sid),
        'questions': data, 'users': data,
        'pno':pno, 'cid':cid, 'num':num, 's_id':s_id, "complete_url":complete_url
        ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/add_question_in_student_section.html", context)




def DeleteMainExam(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = MainExam.objects.filter(id = eid).first()
    if request.user.is_staff:
        data.delete()
    return redirect(request.META.get('HTTP_REFERER'))




def CreateQuestionSection(request, cid, sid, option):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    cour = Sub_course.objects.filter(id = cid).first()
    if option == "List":
        if request.method == "POST":
            name = request.POST['name']
            Question_Section.objects.create(course = cour, name = name)
            return redirect("ExamPanel:create_question_section", cid, sid, option)
    if option == "Delete":
        Question_Section.objects.filter(id = sid).first().delete()
        return redirect("ExamPanel:create_question_section", cid, '0', 'List')
    section = ''
    if option == "Edit":
        section = Question_Section.objects.filter(id=sid).first()
        if request.method == "POST":
            name = request.POST['name']
            section.name = name
            section.save()
            return redirect("ExamPanel:create_question_section", cid, '0', 'List')

    sections = Question_Section.objects.filter(course = cour)
    arg = {
        'course': course, "icour": int(cid), "cour": cid, 'sid':sid,
        'sections':sections, 'section':section, 'option':option,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/CreateQuestionSection.html", arg)


def ExamInstructionView(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = MainExam.objects.filter(id=eid).first()
    return render(request, "exam/exam_instruction.html", {"data":data,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def AddImportQuestions(request, cour, pack, subpack, exam):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    print(type(cour))
    course = Sub_course.objects.all()
    pck = []
    sub = []
    exams = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        if pack != '0':
            pc = ExamPackage.objects.filter(id = pack).first()
            sub = ExamSubPackage.objects.filter(package = pc)
            if subpack != '0':
                sb = ExamSubPackage.objects.filter(id = subpack).first()
                exams = MainExam.objects.filter(Sub_package = sb)

    context = {
        'course':course, "icour":int(cour), "cour":cour,
        'package':pck, 'ipack':int(pack), 'pack':pack,
        'subpackage':sub, 'isubpack':int(subpack), 'subpack':subpack,
        'exams':exams, 'iexam':int(exam), 'exam':exam,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/add_import_questions.html", context)


def AddMainExamSection(request, cour, pack, subpack, exam):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    course = Sub_course.objects.all()
    pck = []
    sub = []
    exams = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        if pack != '0':
            pc = ExamPackage.objects.filter(id = pack).first()
            sub = Package_Subpackage.objects.filter (package = pc)
            if subpack != '0':
                sb = ExamSubPackage.objects.filter(id = subpack).first()
                exams = SubPackage_MainExam.objects.filter(sub = sb)
                if exam != '0':
                    return redirect('ExamPanel:exam_section', cour, pack, subpack, exam)

    context = {
        'course':course, "icour":int(cour), "cour":cour,
        'package':pck, 'ipack':int(pack), 'pack':pack,
        'subpackage':sub, 'isubpack':int(subpack), 'subpack':subpack,
        'exams':exams, 'iexam':int(exam), 'exam':exam,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/add_main_exam_section.html", context)


def ExamSectionCreation(request, cour, pack, subpack, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if eid == '0':
        return redirect("ExamPanel:add_main_exam_section", cour, pack, subpack, eid)
    exam = MainExam.objects.filter(id = eid).first()
    sections = MainExamSection.objects.filter(exam = exam)
    if request.method == "POST":
        d = request.POST
        s = []
        for i in range(1, 6):
            if d['s'+ str(i)]:
                s.append(d['s'+str(i)])
        exam = MainExam.objects.filter(id = eid).first()
        for i in s:
            MainExamSection.objects.create(exam = exam, name = i)
    context = {
        "cour": cour, 'sections':sections,
        'pack': pack,
        'subpack': subpack,
        'exam': eid,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
        }
    return render(request, "exam/exam_sections.html", context)


def ExamSectionDelete(request, sid, cour, pack, subpack, eid ):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = MainExamSection.objects.filter(id = sid).first()
    if request.user.is_staff:
        data.delete()
    return redirect("ExamPanel:exam_section", cour, pack, subpack, eid )

'''

def AllQuestions(request, cour, pack, subpack, exam):
    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(cour) + "__" + str(pack) + "__" + str(subpack) + "__" + str(exam)
    print(type(cour))
    course = Sub_course.objects.all()
    pck = []
    sub = []
    exams = []
    if cour != "0":
        cou = Sub_course.objects.filter(id = cour).first()
        pck = ExamPackage.objects.filter(course=cou)
        if pack != '0':
            pc = ExamPackage.objects.filter(id = pack).first()
            sub = ExamSubPackage.objects.filter(package = pc)
            if subpack != '0':
                sb = ExamSubPackage.objects.filter(id = subpack).first()
                exams = MainExam.objects.filter(Sub_package = sb)

    context = {
        'course':course, "icour":int(cour), "cour":cour,
        'package':pck, 'ipack':int(pack), 'pack':pack,
        'subpackage':sub, 'isubpack':int(subpack), 'subpack':subpack,
        'exams':exams, 'iexam':int(exam), 'exam':exam, "complete_url":complete_url
    }
    return render(request, "exam/all_questions.html", context)






def FilteredQuestions(request, cid, pid, spid, eid):
    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(cid) + "__" + str(pid) + "__" + str(spid) + "__" + str(eid)
    iterator = itertools.count()
    if cid == '0':
        return redirect("ExamPanel:all_questions", '0', '0', '0', '0')
    elif pid == '0':
        ques_tion = []
        d = Sub_course.objects.filter(id = cid).first()
        que = ExamPackage.objects.filter(course = d)
        for pack in que:
            for sub in pack.examsubpackage_set.all():
                for exam in sub.mainexam_set.all():
                    for ques in exam.multiplechoicequestions_set.all():
                        ques_tion.append(ques)

        return render(request, "exam/exam_questions.html", {'questions': ques_tion, 'iterator':iterator, "complete_url":complete_url})

    elif spid == '0':
        ques_tion = []
        d = ExamPackage.objects.filter(id = pid).first()
        subpack = ExamSubPackage.objects.filter(package = d)
        for sub in subpack:
            for exam in sub.mainexam_set.all():
                for ques in exam.multiplechoicequestions_set.all():
                    ques_tion.append(ques)
        return render(request, "exam/exam_questions.html", {'questions': ques_tion, 'iterator':iterator, "complete_url":complete_url})

    elif eid == '0':
        ques_tion = []
        d = ExamSubPackage.objects.filter(id = spid).first()
        que = MainExam.objects.filter(Sub_package = d)
        for exam in que:
            for ques in exam.multiplechoicequestions_set.all():
                ques_tion.append(ques)
        return render(request, "exam/exam_questions.html", {'questions': ques_tion, 'iterator':iterator, "complete_url":complete_url})

    else:
        ques_tion = []
        d = MainExam.objects.filter(id = eid).first()
        que = MultipleChoiceQuestions.objects.filter(exam = d)
        for ques in que:
            ques_tion.append(ques)
        return render(request, "exam/exam_questions.html", {'questions': ques_tion, 'iterator':iterator, "complete_url":complete_url})
'''


def TypeofQuestion(request, sid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if request.method == "POST":
        Type = request.POST['questionType']
        if Type == "0":
            return redirect('ExamPanel:select_question_type', sid)
        if Type == "multi":
            return redirect("ExamPanel:add_multi_questions", sid)
        if Type == "response":
            return redirect("ExamPanel:add_multi_response_questions", sid)
        if Type == "blank":
            return redirect("ExamPanel:add_fill_questions", sid)
        if Type == "truefalse":
            return redirect("ExamPanel:add_true_false_questions", sid)
    return render(request, "exam/questionType.html",{"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def AddMultiQuestions(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = Question_Section.objects.filter(id=eid).first()
    form = MultiChoiseQuestionForm()
    if request.method == "POST":
        form = MultiChoiseQuestionForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.section = d
            data.save()

            return redirect("ExamPanel:add_multi_questions", eid)
    return render(request, "exam/multiquestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def EditMultiQuestions(request, eid, c_url, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = []
    if what == "All":
        data = MultipleChoiceQuestions.objects.filter(id = eid).first()
    if what == "Student":
        data = StudentQuestions.objects.filter(id = eid).first()
        print(data, "Hello")

    form = MultiChoiseQuestionForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        c_url = c_url.split("__")
        name = "ExamPanel:" + c_url.pop(0)
        c_url = tuple(c_url)
        return HttpResponseRedirect(reverse(name, args=c_url))
    return render(request, "exam/edit_multiquestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def AddMultiResponseQuestions(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = Question_Section.objects.filter(id=eid).first()
    form = MultiResponseQuestionForm()
    if request.method == "POST":
        form = MultiResponseQuestionForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.section = d
            data.save()
            return redirect("ExamPanel:add_multi_response_questions", eid)
    return render(request, "exam/multiresponsequestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditMultiResponseQuestions(request, eid, c_url, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = []
    if what == "All":
        data = MultipleChoiceQuestions.objects.filter(id = eid).first()
    if what == "Student":
        data = StudentQuestions.objects.filter(id = eid).first()
    form = MultiResponseQuestionForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        c_url = c_url.split("__")
        name = "ExamPanel:" + c_url.pop(0)
        c_url = tuple(c_url)
        return HttpResponseRedirect(reverse(name, args=c_url))
    return render(request, "exam/edit_multiresponsequestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})




def AddFillInTheBlanksQuestions(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = Question_Section.objects.filter(id=eid).first()
    form = FillInTheBlankQuestionForm()
    if request.method == "POST":
        form = FillInTheBlankQuestionForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.section = d
            data.save()
            return redirect("ExamPanel:add_fill_questions", eid)
    return render(request, "exam/Fillquestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def EditFillInTheBlanksQuestions(request, eid, c_url, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = []
    if what == "All":
        data = MultipleChoiceQuestions.objects.filter(id = eid).first()
    if what == "Student":
        data = StudentQuestions.objects.filter(id = eid).first()
    form = FillInTheBlankQuestionForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        c_url = c_url.split("__")
        name = "ExamPanel:" + c_url.pop(0)
        c_url = tuple(c_url)
        return HttpResponseRedirect(reverse(name, args=c_url))
    return render(request, "exam/edit_Fillquestion.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def AddTrueFalseQuestions(request, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    d = Question_Section.objects.filter(id=eid).first()
    form = TrueFalseQuestionForm()
    if request.method == "POST":
        form = TrueFalseQuestionForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.section = d
            data.save()
            return redirect("ExamPanel:add_fill_questions", eid)
    return render(request, "exam/truefalse.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def EditTrueFalseQuestions(request, eid, c_url, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = []
    if what == "All":
        data = MultipleChoiceQuestions.objects.filter(id = eid).first()
    if what == "Student":
        data = StudentQuestions.objects.filter(id = eid).first()
    form = TrueFalseQuestionForm(request.POST or None, request.FILES or None, instance=data)
    if form.is_valid():
        data = form.save(commit=False)
        data.save()
        c_url = c_url.split("__")
        name = "ExamPanel:" + c_url.pop(0)
        c_url = tuple(c_url)
        return HttpResponseRedirect(reverse(name, args=c_url))

    return render(request, "exam/edit_truefalse.html", {"form":form,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def All_Questions(request, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    course = []
    data = ques_tion
    if cid  != "0":
        course = Sub_course.objects.filter(id = cid).first()
        question_section = Question_Section.objects.filter(course = course)
        for section in question_section:
            for ques in section.multiplechoicequestions_set.all():
                ques_tion.append(ques)
    num = 0
    if cid == '0' and sid != '0':
        section = Question_Section.objects.filter(id=sid).first()
        ques_tion = MultipleChoiceQuestions.objects.filter(section = section)
        page = request.GET.get('page', 1)
        num = int(pno) * (int(page) - 1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)
    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'questions': data, 'users': data,
        'pno': pno, 'cid': sid, 'num': num, "complete_url":complete_url
        ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()
    }
    return render(request, "exam/exam_questions.html", context)




import openpyxl
def Import_Question(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Registered Student List.csv"'
    writer = csv.writer(response)

    if request.method == "POST":
        file = request.FILES['questions']
        print(file.name)
        sheet = request.POST['sheet']
        wb = openpyxl.load_workbook(file, keep_vba=True)
        worksheet = wb[sheet]
        nn = 0
        for row in worksheet.iter_rows():
            n = 0
            main_data = []
            wrong_data = []
            for ceil in row:
                if ceil.value != None:
                    main_data.append(ceil.value)
                wrong_data.append(ceil.value)


            data = list(main_data)
            sec = Question_Section.objects.filter(id = data[0]).first()
            if not sec:
                nn += 1
            if sec:
                if data[1] == "MultiChoice":
                    type1 = type(data[8])
                    type2 = type(data[9])

                    if len(data) != 11 or (type1 != int and type1 != float) or \
                            (type2 != int and type2 != float):
                        nn += 1
                        y = wrong_data.count(None)
                        for i in range(y):
                            z = wrong_data.index(None)
                            wrong_data[z] = " "
                        writer.writerow(wrong_data)
                    else:
                        c = MultipleChoiceQuestions()
                        c.section = sec
                        c.Type1 = True
                        c.Questions = data[2]
                        c.Option_1 = data[3]
                        c.Option_2 = data[4]
                        c.Option_3 = data[5]
                        c.Option_4 = data[6]
                        c.TrueAns = data[7]
                        c.marks1 = data[8]
                        c.minus1 = data[9]
                        c.explain1 = data[10]
                        c.save()

            data = list(main_data)
            sec = Question_Section.objects.filter(id=data[0]).first()
            if sec:
                if data[1] == "MultiResponse":
                    type1 = type(data[11])
                    type2 = type(data[12])

                    Atype1 = type(data[7])
                    Atype2 = type(data[8])
                    Atype3 = type(data[9])
                    Atype4 = type(data[10])

                    if len(data) != 14 or (type1 != int and type1 != float) or \
                            (type2 != int and type2 != float) or Atype1 != int or Atype2 != int \
                            or Atype3 != int or Atype4 != int:
                        y = wrong_data.count(None)
                        for i in range(y):
                            z = wrong_data.index(None)
                            wrong_data[z] = " "
                        writer.writerow(wrong_data)
                        nn += 1
                    else:
                        c = MultipleChoiceQuestions()
                        c.section = sec
                        c.Type2 = True
                        c.Question = data[2]
                        c.Option1 = data[3]
                        c.Option2 = data[4]
                        c.Option3 = data[5]
                        c.Option4 = data[6]
                        c.Answer1 = data[7]
                        c.Answer2 = data[8]
                        c.Answer3 = data[9]
                        c.Answer4 = data[10]
                        c.marks2 = data[11]
                        c.minus2 = data[12]
                        c.explain2 = data[13]
                        c.save()

            data = list(main_data)
            sec = Question_Section.objects.filter(id=data[0]).first()
            if sec:
                if data[1] == "FillBlank":

                    type1 = type(data[5])
                    type2 = type(data[6])

                    if len(data) != 8 or (type1 != int and type1 != float) or \
                            (type2 != int and type2 != float):
                        y = wrong_data.count(None)
                        for i in range(y):
                            z = wrong_data.index(None)
                            wrong_data[z] = " "
                        writer.writerow(wrong_data)
                        nn += 1
                    else:
                        c = MultipleChoiceQuestions()
                        c.section = sec
                        c.Type3 = True
                        c.Que = data[2]
                        c.Answer31 = data[3]
                        if data[4] == "*":
                            c.Answer32 = " "
                        else:
                            c.Answer32 = data[4]
                        c.marks3 = data[5]
                        c.minus3 = data[6]
                        c.explain3 = data[7]
                        c.save()

            data = list(main_data)
            sec = Question_Section.objects.filter(id=data[0]).first()
            if sec:
                if data[1] == "TrueFalse":
                    type1 = type(data[4])
                    type2 = type(data[5])

                    Atype1 = type(data[3])

                    if len(data) != 7 or (type1 != int and type1 != float) or \
                            (type2 != int and type2 != float) or Atype1 != int:
                        y = wrong_data.count(None)
                        for i in range(y):
                            z = wrong_data.index(None)
                            wrong_data[z] = " "
                        writer.writerow(wrong_data)
                        nn += 1
                    else:
                        c = MultipleChoiceQuestions()
                        c.section = sec
                        c.Type4 = True
                        c.Ques_tion = data[2]
                        c.TrueAns1 = data[3]
                        c.marks4 = data[4]
                        c.minus4 = data[5]
                        c.explain4 = data[6]
                        c.save()
        if nn>0:
            return response

        return redirect("ExamPanel:import")
    return render(request, "exam/Import_question_file.html",{"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})



def export_users_xls(request):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"Course Name", 15),
        (u"Section Name", 15),
        (u"Section ID", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

    rows = Question_Section.objects.all()
    for obj in rows:
        row_num += 1
        row = [
            obj.course.name,
            obj.name,
            obj.id,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    wb.save(response)

    ws = wb.get_active_sheet()
    ws.title = "Sections"
    row_num = 0

    columns = [
        (u"Course Name", 15),
        (u"Section Name", 15),
        (u"Section ID", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

    rows = Question_Section.objects.all()
    for obj in rows:
        row_num += 1
        row = [
            obj.course.name,
            obj.name,
            obj.id,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response


def MyQuestionBank(request, cid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(cid) + "__" + str(pno)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        section = Question_Section.objects.filter(course = cour)
        for sec in section:
            for ques in sec.multiplechoicequestions_set.all():
                ques_tion.append(ques)
        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'questions': data, 'users': data,
        'pno':pno, 'cid':cid, 'num':num, "complete_url":complete_url,
        "ins_data": Institute_data(), "adminprofile":Admin_data()
    }
    return render(request, "exam/question_bank.html", context)


def NullQuestion(request, pno, sid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(pno) + "__" + str(sid)
    sections = Question_Section.objects.all()
    ques_tion = MultipleChoiceQuestions.objects.filter(section = None)
    page = request.GET.get('page', 1)
    num = int(pno) * (int(page) - 1)
    if request.method == "POST":
        qu = request.POST.getlist('que')
        s_id = request.POST['sec']
        sec = Question_Section.objects.filter(id = s_id).first()
        for i in qu:
            q = MultipleChoiceQuestions.objects.filter(id=i).first()
            q.section = sec
            q.save()
        return redirect("ExamPanel:null_questions", '10', '0')
    paginator = Paginator(ques_tion, int(pno))
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    context = {
        'questions': data, 'users': data,
        'pno': pno, 'num': num, 'sid':sid, 'sections':sections,
        "complete_url":complete_url,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()

    }
    return render(request, "exam/null_questions.html", context)


def AddQuestion_In_ExamSection(request, s_id, cid, sid, pno):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(s_id) + "__" + str(cid) + "__" + str(sid) + "__" + str(pno)
    ques_tion = []
    data = []
    sections = []
    course = Sub_course.objects.all()
    num = 0
    if cid != "0":
        cour = Sub_course.objects.filter(id=cid).first()
        sections = Question_Section.objects.filter(course = cour)
        if sid != "0":
            section = Question_Section.objects.filter(id=sid).first()
            all_ques_tion = MultipleChoiceQuestions.objects.filter(section = section)

            sec = MainExamSection.objects.filter(id=s_id).first()
            Exam = MainExam.objects.filter(id = sec.exam.id).first()
            all_sec = MainExamSection.objects.filter(exam = Exam)
            pre_que = ExamQuestion.objects.filter(MainSection=sec)

            for i in all_ques_tion:
                n = 0
                for z in all_sec:
                    pre_que = ExamQuestion.objects.filter(MainSection=z)
                    for j in pre_que:
                        if i.id == j.Question.id:
                            n+=1
                if n == 0:
                    ques_tion.append(i)

            page = request.GET.get('page', 1)
            num = int(pno) * (int(page)-1)
            paginator = Paginator(ques_tion, int(pno))
            try:
                data = paginator.page(page)
            except PageNotAnInteger:
                data = paginator.page(1)
            except EmptyPage:
                data = paginator.page(paginator.num_pages)

            if request.method == "POST":
                print("Hello")
                qu = request.POST.getlist('que')
                sec = MainExamSection.objects.filter(id=s_id).first()
                for i in qu:
                    que = MultipleChoiceQuestions.objects.filter(id = i).first()
                    if que:
                        ExamQuestion.objects.create(MainSection = sec, Question = que)
                return redirect("ExamPanel:add_questions_in_exam_section", s_id, cid, sid, pno)


    context = {
        'course': course, "icour": int(cid), "cour": cid,
        'sections':sections, 'sec':sid, "isec":int(sid),
        'questions': data, 'users': data,"adminprofile":Admin_data(),
        'pno':pno, 'cid':cid, 'num':num, 's_id':s_id, "complete_url":complete_url,
        "ins_data": Institute_data()
    }
    return render(request, "exam/add_question_in_section.html", context)



def ExamSectionQuestionList(request, sid, pno, what):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    current_url = resolve(request.path_info).url_name
    complete_url = str(current_url) + "__" + str(sid) + "__" + str(pno) + "__" + str(what)
    ques_tion = []
    data = []
    course = Sub_course.objects.all()
    num = 0
    if sid != "0":
        if what == "All":
            exam = MainExam.objects.filter(id = sid).first()
            sections = MainExamSection.objects.filter(exam = exam)
            for s in sections:
                all_eq = ExamQuestion.objects.filter(MainSection=s)
                for i in all_eq:
                    ques_tion.append(i)
        else:
            sec = MainExamSection.objects.filter(id = sid).first()
            all_eq = ExamQuestion.objects.filter(MainSection = sec)
            for i in all_eq:
                ques_tion.append(i)

        page = request.GET.get('page', 1)
        num = int(pno) * (int(page)-1)
        paginator = Paginator(ques_tion, int(pno))
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            data = paginator.page(1)
        except EmptyPage:
            data = paginator.page(paginator.num_pages)

    context = {
        'questions': data, 'users': data,"adminprofile":Admin_data(),
        'pno':pno, 'num':num, 'cid':sid, "what":what, "complete_url":complete_url,
        "ins_data": Institute_data()
    }
    return render(request, "exam/ExamSectionQuestion.html", context)



def RemoveQuestions(request, qid, where):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    if where == "exam":
        data = ExamQuestion.objects.filter(id = qid).first()
        if request.user.is_staff:
            data.delete()
    if where == "quiz":
        data = Quiz_Questions.objects.filter(id=qid).first()
        if request.user.is_staff:
            data.delete()
    if where == "study":
        data = Topic_Question.objects.filter(id=qid).first()
        if request.user.is_staff:
            data.delete()
    if where == "all":
        data = MultipleChoiceQuestions.objects.filter(id=qid).first()
        if request.user.is_staff:
            data.delete()
    if where == "Student":
        data = StudentExamQuestion.objects.filter(id=qid).first()
        if request.user.is_staff:
            data.delete()
    if where == "studentquiz":
        data = StudentQuiz_Questions.objects.filter(id=qid).first()
        if request.user.is_staff:
            data.delete()
    return redirect(request.META.get('HTTP_REFERER'))


def Exam_generate_invoice(request,pid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    data = ExamInvoiceGenerate.objects.filter(id = pid).first()

    return render(request,'exam/exam_generate_invoice.html',{"data":data
                                                               ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data()})


def Publish_UnPublish(request, eid, what, what1):
    if what == "Exam":
        exam = ExamPackage.objects.filter(id = eid).first()
        if what1 == "Unpublish":
            exam.status = "UnPublish"
            exam.save()
            return redirect(request.META.get('HTTP_REFERER'))
        if what1 == "Publish":
            questions = []
            status = False
            for sub in exam.package_subpackage_set.all():
                sub_pack = ExamSubPackage.objects.filter(id = sub.sub.id).first()
                for main in sub_pack.subpackage_mainexam_set.all():
                    Main = MainExam.objects.filter(id = main.main.id).first()
                    for Sec in Main.mainexamsection_set.all():
                        if Sec.examquestion_set.first():
                            status = True
                        else:
                            status = False
            if status:
                exam.status = "Publish"
                exam.save()
        return redirect(request.META.get('HTTP_REFERER'))
    if what == "Quiz":
        quiz = QuizPackage.objects.filter(id=eid).first()
        if what1 == "Unpublish":
            quiz.status = "UnPublish"
            quiz.save()
        if what1 == "Publish":
            print("Hello")
            questions1 = []
            status = False
            for qui in quiz.quiz_quizpackage_set.all():
                status = False
                qu = Quiz.objects.filter(id = qui.quiz.id).first()
                print("Hello")
                for que in qu.quiz_questions_set.all():
                    questions1.append(que)
                    status = True

            if status:
                print("Hello")
                quiz.status = "Publish"
                quiz.save()
        return redirect(request.META.get('HTTP_REFERER'))




def Test_Package_Result(request,cor,bat, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = ExamPackage.objects.filter(status = "Publish")
    students = False
    batch = False
    course = False
    SubPacks = []
    assigns = []
    total_marks = 0
    n = 0
    Exams = []
    if cor != '0':
        course = ExamPackage.objects.filter(id= cor).first()
        SubPackage = course.package_subpackage_set.all()
        SubPacks = []
        for i in SubPackage:
            sub = ExamSubPackage.objects.filter(id = i.sub.id).first()
            SubPacks.append(sub)
    if bat != '0':
        batch = ExamSubPackage.objects.filter(id = bat).first()
        Exam = batch.subpackage_mainexam_set.all()
        for i in Exam:
            main = MainExam.objects.filter(id = i.main.id).first()
            Exams.append(main)
    if eid != '0':
        main = MainExam.objects.filter(id=eid).first()
        assigns = StudentMainExam.objects.filter(main = main, status = "Submitted")
        n = 0
        total_marks = 0
        for sec in main.mainexamsection_set.all():
            n += len(sec.examquestion_set.all())

            que = ExamQuestion.objects.filter(MainSection=sec)
            for i in que:
                t = 0
                f = 0
                if i.Question.Type1:
                    t = i.Question.marks1
                elif i.Question.Type2:
                    t = i.Question.marks2
                elif i.Question.Type3:
                    t = i.Question.marks3
                elif i.Question.Type4:
                    t = i.Question.marks4
                if t:
                    total_marks += float(t)
    context = {"all_course":all_course, "students":assigns, "subpackages":SubPacks, 'pid':cor, 'ipid':int(cor),
               'isid':int(bat), 'sid':bat, 'eid':eid, 'ieid':int(eid), 'total':total_marks, 'ques':n
               ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data(), 'Exams':Exams}
    return render(request,'Practice/test_package_result.html',context)


def Quiz_Package_Result(request,cor,bat, eid):
    if not request.user.is_authenticated():
        return redirect("login")
    user = request.user
    if not is_Examinar(user) and not is_admin(user):
        return redirect("login")

    all_course = QuizPackage.objects.filter(status = "Publish")
    students = False
    batch = False
    course = False
    SubPacks = []
    assigns = []
    total_marks = 0
    n = 0
    Exams = []
    if cor != '0':
        course = QuizPackage.objects.filter(id= cor).first()
        quizs = course.quiz_quizpackage_set.all()
        for i in quizs:
            sub = Quiz.objects.filter(id = i.quiz.id).first()
            SubPacks.append(sub)
    if bat != '0':
        main = Quiz.objects.filter(id=bat).first()
        print(main)
        assigns = StudentQuiz.objects.filter(main = main, status = "Submitted")
        n = len(main.quiz_questions_set.all())
        total_marks = 0
        for i in main.quiz_questions_set.all():
            if i.Question.Type1:
                t = i.Question.marks1
            elif i.Question.Type2:
                t = i.Question.marks2
            elif i.Question.Type3:
                t = i.Question.marks3
            elif i.Question.Type4:
                t = i.Question.marks4
            if t:
                total_marks += float(t)
    context = {"all_course":all_course, "students":assigns, "subpackages":SubPacks, 'pid':cor, 'ipid':int(cor),
               'isid':int(bat), 'sid':bat, 'eid':eid, 'ieid':int(eid), 'total':total_marks, 'ques':n
               ,"ins_data":Institute_data(request.user), "adminprofile":Admin_data(), 'Exams':Exams}
    return render(request,'Practice/quiz_package_result.html',context)
